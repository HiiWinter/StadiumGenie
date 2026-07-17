import os
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load Algeria player JSON
with open("algeria_players.json", "r", encoding="utf-8") as f:
    algeria_players = json.load(f)

with open("algeria_helper_map.json", "r", encoding="utf-8") as f:
    algeria_helper_map = json.load(f)

# 1. Update src/utils/playerImageHelper.js
helper_path = r"e:\Challenge for\src\utils\playerImageHelper.js"
with open(helper_path, "r", encoding="utf-8") as f:
    helper_code = f.read()

# Add Algeria mappings at top of localPlayerFiles object
entries_str = ""
for k, v in algeria_helper_map.items():
    entries_str += f'  "{k}": "{v}",\n'

if '"alg_melvin_mastil.jpg"' not in helper_code:
    helper_code = helper_code.replace("const localPlayerFiles = {\n", "const localPlayerFiles = {\n" + entries_str)
    with open(helper_path, "w", encoding="utf-8") as f:
        f.write(helper_code)
    print("Updated playerImageHelper.js with Algeria entries.")

# 2. Update src/components/SpectatorHub.jsx
spectator_path = r"e:\Challenge for\src\components\SpectatorHub.jsx"
with open(spectator_path, "r", encoding="utf-8") as f:
    spectator_code = f.read()

alg_formatted_js = ""
for p in algeria_players:
    alg_formatted_js += "  {\n"
    alg_formatted_js += f'    "name": "{p["name"]}",\n'
    alg_formatted_js += f'    "code": "{p["code"]}",\n'
    alg_formatted_js += f'    "country": "{p["country"]}",\n'
    alg_formatted_js += f'    "flag": "{p["flag"]}",\n'
    alg_formatted_js += f'    "position": "{p["position"]}",\n'
    alg_formatted_js += f'    "height": "{p["height"]}",\n'
    alg_formatted_js += f'    "weight": "{p["weight"]}",\n'
    alg_formatted_js += f'    "age": {p["age"]},\n'
    alg_formatted_js += f'    "club": "{p["club"]}",\n'
    alg_formatted_js += f'    "league": "{p["league"]}",\n'
    alg_formatted_js += f'    "goals": {p["goals"]},\n'
    alg_formatted_js += f'    "caps": {p["caps"]},\n'
    alg_formatted_js += f'    "marketValue": "{p["marketValue"]}",\n'
    alg_formatted_js += f'    "image": "{p["image"]}"\n'
    alg_formatted_js += "  },\n"

if '"Melvin MASTIL"' not in spectator_code:
    spectator_code = spectator_code.replace("const PLAYER_DATABASE = [\n", "const PLAYER_DATABASE = [\n" + alg_formatted_js)
    
    # Check if Algeria is in QUALIFIED_TEAMS array in SpectatorHub.jsx
    if 'code: "ALG"' not in spectator_code and 'code: \'ALG\'' not in spectator_code:
        alg_team_obj = """  {
    code: 'ALG',
    name: 'Algeria',
    flag: '🇩🇿',
    confederation: 'CAF',
    group: 'Group F',
    fifaRank: 37,
    keyPlayers: ['Riyad Mahrez', 'Rayan Aït-Nouri', 'Amine Gouiri'],
    manager: 'Vladimir Petkovic',
    topScorer: 'Riyad Mahrez (31 goals)',
    color: '#006233'
  },\n"""
        spectator_code = spectator_code.replace("const QUALIFIED_TEAMS = [\n", "const QUALIFIED_TEAMS = [\n" + alg_team_obj)
        print("Added Algeria to QUALIFIED_TEAMS array in SpectatorHub.jsx")

    with open(spectator_path, "w", encoding="utf-8") as f:
        f.write(spectator_code)
    print("Updated SpectatorHub.jsx with Algeria squad.")

# 3. Extract ALL players from SpectatorHub.jsx to create complete Excel file
match = re.search(r'const PLAYER_DATABASE = \[\s*([\s\S]*?)\n\];', spectator_code)
if match:
    raw_array = "[" + match.group(1) + "]"
    # Fix trailing commas if any
    raw_array = re.sub(r',\s*\]', ']', raw_array)
    all_players = json.loads(raw_array)
    print(f"Extracted {len(all_players)} total players from database for Excel export.")
else:
    print("Error: Could not parse PLAYER_DATABASE array from SpectatorHub.jsx")
    all_players = []

# Generate Excel Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FIFA 2026 Player Database"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# Colors & Styling
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark slate
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

headers = [
    "No", "Player Name", "Country", "Code", "Flag", "Position", 
    "Age", "Height", "Weight", "Club", "League", 
    "Caps", "Goals", "Market Value", "Image Path"
]

ws.append(headers)

# Apply header styles
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center" if col_num in [1, 4, 5, 7, 8, 9, 12, 13] else "left", vertical="center")
    cell.border = thin_border

ws.row_dimensions[1].height = 26

# Populate Data Rows
for idx, p in enumerate(all_players, 1):
    row_data = [
        idx,
        p.get("name", ""),
        p.get("country", ""),
        p.get("code", ""),
        p.get("flag", ""),
        p.get("position", ""),
        p.get("age", ""),
        p.get("height", ""),
        p.get("weight", ""),
        p.get("club", ""),
        p.get("league", ""),
        p.get("caps", 0),
        p.get("goals", 0),
        p.get("marketValue", ""),
        p.get("image", "")
    ]
    ws.append(row_data)
    row_num = idx + 1
    fill = zebra_fill if idx % 2 == 0 else white_fill
    ws.row_dimensions[row_num].height = 20
    
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = Font(name="Segoe UI", size=10)
        cell.border = thin_border
        
        # Alignment
        if c in [1, 4, 5, 7, 12, 13]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c in [8, 9, 14]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Auto-fit columns with padding
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

excel_filename = "StadiumGenie_Player_Database_2026.xlsx"
wb.save(excel_filename)
wb.save(os.path.join("public", excel_filename))
print(f"Successfully created Excel file: {excel_filename}")
